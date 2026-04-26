"""
KHK/SSMS 导出的 `Odoo_Vendor_Import.xlsx` → 清洗后 `Odoo_Vendor_Import_clean.xlsx`。
**列 `ref` = KHK 供应商编号（如 70707）**，与供应商价目表里的 Vendor (ref) 一致，必须在 Odoo 里导入为联系人的 **内部参考 (Internal Reference)**。

若已手工建商、无 ref：用本文件重导并选「按名称更新」+ 映射 ref；或联系人列表导出后补一列 ref 再导入更新。
`ref` 在表单上常在 **销售与采购 (Sales & Purchase)** 页签，不在主联系人区。
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_FILE = 'Odoo_Vendor_Import.xlsx'
OUTPUT_FILE = 'Odoo_Vendor_Import_clean.xlsx'

df = pd.read_excel(INPUT_FILE)

# Clean NULL strings and whitespace
for col in df.columns:
    if df[col].dtype == object:
        df[col] = df[col].str.strip()
        df[col] = df[col].replace({'NULL': None, 'null': None, 'nan': None, '': None})

# Build Odoo import format
# Column names match Odoo fields directly — no manual mapping needed on import
odoo = pd.DataFrame()
odoo['ref']                  = df['Vendor_ID'].astype(str)
odoo['name']                 = df['Company_Name']
odoo['is_company']           = True
# Odoo: partner is a **Vendor** when supplier_rank > 0 (without this, only a contact, not a supplier)
odoo['supplier_rank']        = 1
odoo['active']               = True
odoo['street']               = df['Street']
odoo['zip']                  = df['Zip_Code'].astype(str).str.strip().where(df['Zip_Code'].notna(), None)
odoo['city']                 = df['City']
odoo['country_id']           = df['Country']
odoo['phone']                = df['Phone']
odoo['email']                = df['Email']
odoo['bank_ids/acc_number']  = df['IBAN']
odoo['bank_ids/bank_bic']    = df['BIC_Swift']

odoo.to_excel(OUTPUT_FILE, index=False)

# Styling
wb = load_workbook(OUTPUT_FILE)
ws = wb.active

header_fill = PatternFill('solid', start_color='1F4E79', end_color='1F4E79')
header_font = Font(name='Arial', bold=True, color='FFFFFF', size=11)
border = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF')
)

col_widths = [12, 35, 12, 8, 25, 10, 15, 10, 15, 25, 30, 15]
for i, (cell, w) in enumerate(zip(ws[1], col_widths), 1):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    ws.column_dimensions[get_column_letter(i)].width = w

data_font = Font(name='Arial', size=10)
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.font = data_font
        cell.border = border
        cell.alignment = Alignment(vertical='center')

ws.freeze_panes = 'A2'
ws.auto_filter.ref = ws.dimensions
ws.row_dimensions[1].height = 22

wb.save(OUTPUT_FILE)
print(f"Completed！Overall Export {len(odoo)} -> {OUTPUT_FILE}")
