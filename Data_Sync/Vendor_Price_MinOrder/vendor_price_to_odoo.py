"""
Sage 100 -> Odoo Vendor Pricelist Import Script (v2 - ID-based matching)
========================================================================
Reads the exported vendor price data and the two Odoo ID mapping files,
then generates an Excel file with a sheet ready to import into Odoo
product.supplierinfo (Vendor Pricelists) using database IDs.

This avoids all "no matching name" errors by matching on database IDs
instead of display names.

Required input files (place in the same folder as this script):
    1. Vendor_Price_MinOrder.xlsx   - exported from SSMS using export_vendor_price_mssql.sql
    2. odoo_partner_id_map.xlsx     - exported from Odoo Contacts with columns: ID, Reference, Name
    3. odoo_product_tmpl_id_map.xlsx - exported from Odoo Products with columns: ID, Internal Reference, Name

How to export the mapping files from Odoo:
    Contacts -> select all -> gear icon -> Export -> fields: ID, Reference, Name
    Inventory -> Products -> select all -> gear icon -> Export -> fields: ID, Internal Reference, Name

Output file:
    Odoo_Vendor_Pricelists_Import_v2.xlsx
    Sheets:
      - Import_ById   : use this sheet in Odoo import wizard (map by Database ID)
      - Preflight     : lists all vendor refs and product codes in the file,
                        for verifying they all exist in Odoo before importing
      - Unmatched     : rows where vendor or product could not be matched (for debugging)

Import steps in Odoo:
    Purchase -> Configuration -> Vendor Pricelists  (or via any product's Purchase tab)
    Gear icon -> Import -> Load file -> select sheet "Import_ById"
    Field mapping:
      Vendor (Database id)           -> Vendor / Database ID
      Product Template (Database id) -> Product Template / Database ID
      Unit Price                     -> Unit Price
      Minimal Quantity               -> Quantity
      Vendor Product Name            -> Vendor Product Name
      Vendor Product Code            -> Vendor Product Code

Dependencies:
    pip install pandas openpyxl

Usage:
    python vendor_price_to_odoo_v2.py
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

INPUT_PRICE_FILE   = os.path.join(BASE_DIR, "Vendor_Price_MinOrder.xlsx")
PARTNER_MAP_FILE   = os.path.join(BASE_DIR, "odoo_partner_id_map.xlsx")
PRODUCT_MAP_FILE   = os.path.join(BASE_DIR, "odoo_product_tmpl_id_map.xlsx")
OUTPUT_FILE        = os.path.join(BASE_DIR, "Odoo_Vendor_Pricelists_Import_v2.xlsx")

# UoM mapping: Sage 100 unit codes -> Odoo unit names
# Extend this dict if your data contains other units
UOM_MAP = {
    "Stk":   "Units",
    "Stck":  "Units",
    "Stck.": "Units",
    "Ktn":   "Units",
    "Ktn.":  "Units",
    "Pck":   "Units",
    "Pck.":  "Units",
    "Btl":   "Btl",
    "Dose":  "Dose",
    "Rolle": "Rolle",
    "Str":   "Str",
}

# ─────────────────────────────────────────────
# Helper: normalize key for matching
# (handles float->int conversion, e.g. 70707.0 -> "70707")
# ─────────────────────────────────────────────

def normalize_key(value) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    if isinstance(value, bool):
        return ""
    if isinstance(value, float):
        if value == int(value):
            return str(int(value))
        return str(value).strip()
    s = str(value).strip()
    # Convert "70707.0" -> "70707"
    if s.endswith(".0") and s[:-2].lstrip("-").isdigit():
        return s[:-2]
    return s if s.lower() not in ("nan", "none", "") else ""


def normalize_series(series: pd.Series) -> pd.Series:
    return series.map(normalize_key)


# ─────────────────────────────────────────────
# Load input files
# ─────────────────────────────────────────────

def load_files():
    for path in [INPUT_PRICE_FILE, PARTNER_MAP_FILE, PRODUCT_MAP_FILE]:
        if not os.path.exists(path):
            raise FileNotFoundError(
                f"Required file not found: '{path}'\n"
                "Place all three input files in the same folder as this script."
            )

    price_df   = pd.read_excel(INPUT_PRICE_FILE)
    partner_df = pd.read_excel(PARTNER_MAP_FILE)
    product_df = pd.read_excel(PRODUCT_MAP_FILE)

    # Normalize column names (strip whitespace)
    price_df.columns   = [str(c).strip() for c in price_df.columns]
    partner_df.columns = [str(c).strip() for c in partner_df.columns]
    product_df.columns = [str(c).strip() for c in product_df.columns]

    return price_df, partner_df, product_df


# ─────────────────────────────────────────────
# Build lookup dicts: normalized_key -> Odoo database ID
# ─────────────────────────────────────────────

def build_partner_lookup(partner_df: pd.DataFrame) -> dict:
    # Expects columns: ID, Reference, Name
    ref_col = next((c for c in partner_df.columns if "reference" in c.lower()), None)
    id_col  = next((c for c in partner_df.columns if c.lower() == "id"), None)
    if not ref_col or not id_col:
        raise ValueError(f"Partner map missing 'ID' or 'Reference' column. Found: {list(partner_df.columns)}")

    lookup = {}
    for _, row in partner_df.iterrows():
        key = normalize_key(row[ref_col])
        if key:
            lookup[key] = int(row[id_col])
    return lookup


def build_product_lookup(product_df: pd.DataFrame) -> dict:
    # Expects columns: ID, Internal Reference, Name
    ref_col = next((c for c in product_df.columns if "internal" in c.lower() or "reference" in c.lower()), None)
    id_col  = next((c for c in product_df.columns if c.lower() == "id"), None)
    if not ref_col or not id_col:
        raise ValueError(f"Product map missing 'ID' or 'Internal Reference' column. Found: {list(product_df.columns)}")

    lookup = {}
    for _, row in product_df.iterrows():
        key = normalize_key(row[ref_col])
        if key:
            lookup[key] = int(row[id_col])
    return lookup


# ─────────────────────────────────────────────
# Build import DataFrame
# ─────────────────────────────────────────────

def build_import(price_df, partner_lookup, product_lookup):
    rows_ok      = []
    rows_unmatched = []

    for _, row in price_df.iterrows():
        vendor_key  = normalize_key(row.get("Vendor_ID", ""))
        product_key = normalize_key(row.get("Product_Number", ""))
        price       = row.get("Purchase_Price")
        min_qty     = row.get("Min_Order_Qty", 0)
        prod_name   = str(row.get("Product_Name", "")).strip()
        vendor_code = normalize_key(row.get("Vendor_Product_Code", ""))
        uom_raw     = str(row.get("UoM", "Stk")).strip()
        uom         = UOM_MAP.get(uom_raw, uom_raw if uom_raw else "Units")

        # Skip rows without price
        try:
            if pd.isna(price) or float(price) <= 0:
                continue
        except (TypeError, ValueError):
            continue

        partner_id = partner_lookup.get(vendor_key)
        product_id = product_lookup.get(product_key)

        reason = []
        if not partner_id:
            reason.append(f"Vendor ref '{vendor_key}' not found in Odoo")
        if not product_id:
            reason.append(f"Product '{product_key}' not found in Odoo")

        if partner_id and product_id:
            rows_ok.append({
                "Vendor (Database id)":           partner_id,
                "Product Template (Database id)": product_id,
                "Vendor Product Name":            prod_name,
                "Vendor Product Code":            vendor_code,
                "Unit":                           uom,
                "Unit Price":                     float(price),
                "Minimal Quantity":               max(float(min_qty) if pd.notna(min_qty) else 0, 0),
            })
        else:
            rows_unmatched.append({
                "Vendor_ID":      vendor_key,
                "Product_Number": product_key,
                "Product_Name":   prod_name,
                "Purchase_Price": price,
                "Reason":         " | ".join(reason),
            })

    import_df    = pd.DataFrame(rows_ok)
    unmatched_df = pd.DataFrame(rows_unmatched)
    return import_df, unmatched_df


# ─────────────────────────────────────────────
# Build preflight sheet
# ─────────────────────────────────────────────

def build_preflight(price_df, partner_lookup, product_lookup):
    vendor_keys  = sorted(set(normalize_key(v) for v in price_df["Vendor_ID"] if normalize_key(v)))
    product_keys = sorted(set(normalize_key(p) for p in price_df["Product_Number"] if normalize_key(p)))

    vendor_rows = [
        {"Vendor_ref": k, "Found_in_Odoo": "YES" if k in partner_lookup else "NO"}
        for k in vendor_keys
    ]
    product_rows = [
        {"Product_default_code": k, "Found_in_Odoo": "YES" if k in product_lookup else "NO"}
        for k in product_keys
    ]

    max_len = max(len(vendor_rows), len(product_rows))
    data = {
        "Vendor_ref":          [r["Vendor_ref"] for r in vendor_rows] + [""] * (max_len - len(vendor_rows)),
        "Vendor_found":        [r["Found_in_Odoo"] for r in vendor_rows] + [""] * (max_len - len(vendor_rows)),
        "Product_default_code":[r["Product_default_code"] for r in product_rows] + [""] * (max_len - len(product_rows)),
        "Product_found":       [r["Found_in_Odoo"] for r in product_rows] + [""] * (max_len - len(product_rows)),
    }
    return pd.DataFrame(data)


# ─────────────────────────────────────────────
# Style workbook
# ─────────────────────────────────────────────

def style_workbook(output_file):
    wb = load_workbook(output_file)
    header_fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    data_font   = Font(name="Arial", size=10)
    border      = Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )
    col_widths = {
        "Import_ById": [22, 30, 36, 22, 10, 12, 18],
        "Preflight":   [18, 14, 24, 14],
        "Unmatched":   [16, 20, 36, 14, 50],
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        widths = col_widths.get(sheet_name, [])
        for i, cell in enumerate(ws[1], 1):
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border
            w = widths[i - 1] if i <= len(widths) else 16
            ws.column_dimensions[get_column_letter(i)].width = w

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font      = data_font
                cell.border    = border
                cell.alignment = Alignment(vertical="center")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 22

    wb.save(output_file)


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────

def main():
    print("Loading input files...")
    try:
        price_df, partner_df, product_df = load_files()
    except FileNotFoundError as e:
        print(f"ERROR: {e}")
        return

    print(f"  Price data   : {len(price_df)} rows")
    print(f"  Partner map  : {len(partner_df)} rows")
    print(f"  Product map  : {len(product_df)} rows")

    print("\nBuilding lookup tables...")
    try:
        partner_lookup = build_partner_lookup(partner_df)
        product_lookup = build_product_lookup(product_df)
    except ValueError as e:
        print(f"ERROR: {e}")
        return

    print(f"  Partner lookup: {len(partner_lookup)} entries")
    print(f"  Product lookup: {len(product_lookup)} entries")

    print("\nMatching rows...")
    import_df, unmatched_df = build_import(price_df, partner_lookup, product_lookup)
    preflight_df = build_preflight(price_df, partner_lookup, product_lookup)

    # Match rate stats
    total        = len(import_df) + len(unmatched_df)
    match_rate   = len(import_df) / total * 100 if total > 0 else 0
    vendor_miss  = preflight_df[preflight_df["Vendor_found"] == "NO"]["Vendor_ref"].tolist()
    product_miss = preflight_df[preflight_df["Product_found"] == "NO"]["Product_default_code"].tolist()

    print(f"  Matched      : {len(import_df)}/{total} rows ({match_rate:.1f}%)")
    print(f"  Unmatched    : {len(unmatched_df)} rows")
    if vendor_miss:
        print(f"  Missing vendor refs  ({len(vendor_miss)}): {vendor_miss[:5]}{'...' if len(vendor_miss) > 5 else ''}")
    if product_miss:
        print(f"  Missing product codes({len(product_miss)}): {product_miss[:5]}{'...' if len(product_miss) > 5 else ''}")

    print(f"\nWriting output: {OUTPUT_FILE}")
    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        import_df.to_excel(writer,    sheet_name="Import_ById", index=False)
        preflight_df.to_excel(writer, sheet_name="Preflight",   index=False)
        unmatched_df.to_excel(writer, sheet_name="Unmatched",   index=False)

    style_workbook(OUTPUT_FILE)

    print("\nDone!")
    print(f"  Output file  : {OUTPUT_FILE}")
    print()
    print("Import steps in Odoo:")
    print("  Purchase -> Configuration -> Vendor Pricelists (or any product Purchase tab)")
    print("  Gear icon -> Import -> Load file -> select sheet 'Import_ById'")
    print("  Field mapping:")
    print("    Vendor (Database id)           -> Vendor / Database ID")
    print("    Product Template (Database id) -> Product Template / Database ID")
    print("    Unit Price                     -> Unit Price")
    print("    Minimal Quantity               -> Quantity")
    print("    Vendor Product Name            -> Vendor Product Name")
    print("    Vendor Product Code            -> Vendor Product Code")


if __name__ == "__main__":
    main()